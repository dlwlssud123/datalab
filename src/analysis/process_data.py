from pathlib import Path
from datetime import time
from statistics import median
import sys
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Windows 콘솔 한글/이모지 출력 인코딩 설정
sys.stdout.reconfigure(encoding="utf-8")

# 1. 경로 설정
BASE_DIR = Path(__file__).resolve().parent.parent.parent
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"
BUS_SCHEDULE_RULES = {
    "wind_hill": {
        "label": "해금강 정류장(바람의 언덕/신선대 접근권)",
        "fallback_interval_min": 120,
        "fallback_routes_count": 1,
        "sources": [{
            "file": "40번대50번대.xlsx",
            "sheet": "50번대(고현-동부,남부)",
            "direction_keyword": "고현 →",
            "stop_names": ["해금강"],
            "route_numbers": ["55", "55-1"],
        }],
    },
    "maemi_castle": {
        "label": "매미성 직접 정류장",
        "fallback_interval_min": 95,
        "fallback_routes_count": 1,
        "fallback_reason": "엑셀 시간표에서 매미성 직접 정류장명을 찾지 못해 기존 환승 대기 포함 기준값 유지",
        "sources": [{
            "file": "10번대20번대30번대.xlsx",
            "sheet": "32,33번대(고현-두모실,율천-능포)",
            "direction_keyword": "고현 →",
            "stop_names": ["매미성", "복항", "대금"],
        }],
    },
    "hakdong_pebble": {
        "label": "학동 정류장",
        "fallback_interval_min": 80,
        "fallback_routes_count": 2,
        "sources": [{
            "file": "40번대50번대.xlsx",
            "sheet": "50번대(고현-동부,남부)",
            "direction_keyword": "고현 →",
            "stop_names": ["학동"],
            "route_numbers": ["55", "67", "67-1"],
        }],
    },
    "geoje_panoramic": {
        "label": "거제케이블카승강장 경유 55번 계열",
        "fallback_interval_min": 90,
        "fallback_routes_count": 1,
        "sources": [{
            "file": "40번대50번대.xlsx",
            "sheet": "50번대(고현-동부,남부)",
            "direction_keyword": "고현 →",
            "stop_names": ["고현"],
            "note_keywords": ["거제케이블카승강장"],
            "route_numbers": ["55", "55-1"],
        }],
    },
    "geoje_jungle_dome": {
        "label": "식물원/외간교회 정류장",
        "fallback_interval_min": 110,
        "fallback_routes_count": 1,
        "sources": [{
            "file": "40번대50번대.xlsx",
            "sheet": "50번대(고현-동부,남부)",
            "direction_keyword": "고현 →",
            "stop_names": ["식물원", "외간교회"],
            "route_numbers": ["50-2"],
        }],
    },
}

def read_csv_safe(filepath, **kwargs):
    """utf-8, cp949 등 다양한 인코딩을 자동으로 시도하여 안전하게 CSV를 로드합니다."""
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        try:
            return pd.read_csv(filepath, encoding=enc, **kwargs)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return pd.read_csv(filepath, **kwargs)


def load_raw_datasets():
    """올려주신 실제 원천 파일들을 자동으로 찾아 로드합니다."""
    print("[1] 실제 원천 데이터 파일 로드 중...")

    # (1) 인기관광지 (외지인)
    spot_file = list(RAW_DIR.rglob("*인기관광지_외지인.csv"))[0]
    df_spots_rank = read_csv_safe(spot_file)
    print(f"  - 인기관광지 파일 로드 완료: {spot_file.name}")

    # (2) 체류시간 데이터
    stay_file = list(RAW_DIR.rglob("*평균 체류시간 추이(외지인).csv"))[0]
    df_stay = read_csv_safe(stay_file)
    avg_stay_geoje = df_stay[df_stay["지역명"] == "거제시"]["체류시간(분)"].mean()
    print(f"  - 거제시 실측 평균 체류시간: {avg_stay_geoje:.1f}분 (약 {avg_stay_geoje/60:.1f}시간)")

    # (3) 소상공인 상가 데이터 (경남 데이터에서 거제시만 추출)
    store_file = list(RAW_DIR.rglob("*상가(상권)정보_경남*.csv"))[0]
    print(f"  - 소상공인 상가 데이터 로드 중 (용량이 커서 2~3초 소요)...")
    df_stores_all = read_csv_safe(store_file, low_memory=False)
    
    df_geoje_stores = df_stores_all[df_stores_all["시군구명"] == "거제시"]
    print(f"  - 거제시 상가 필터링 완료: 총 {len(df_geoje_stores):,}개 업소")

    return df_spots_rank, avg_stay_geoje, df_geoje_stores



def calculate_entropy_by_town(df_geoje_stores):
    """행정동별 실제 상권 섀넌 엔트로피(CEI) 다양성 지수 계산"""
    print("\n[2] 거제시 읍면동별 실제 상권 엔트로피(CEI) 계산 중...")
    town_entropy = {}

    for town, group in df_geoje_stores.groupby("행정동명"):
        # 업종 중분류별 점유율(p) 계산
        counts = group["상권업종중분류명"].value_counts()
        probs = counts / counts.sum()
        
        # 섀넌 엔트로피 공식: H = - sum(p * log2(p))
        entropy = -np.sum(probs * np.log2(probs))
        
        # 0 ~ 1 사이로 정규화 (이론상 최대 엔트로피 기준)
        max_possible_entropy = np.log2(len(counts)) if len(counts) > 1 else 1.0
        normalized_cei = round(float(entropy / max_possible_entropy), 2)
        town_entropy[town] = normalized_cei

    print("  - 읍면동별 상권 다양성 계산 완료!")
    return town_entropy

def normalize_text(value):
    return re.sub(r"\s+", "", str(value or ""))

def extract_exact_time(value):
    if isinstance(value, time):
        return [value.hour * 60 + value.minute]
    
    if isinstance(value, str) and re.fullmatch(r"\s*([0-2]?\d)[:：]([0-5]\d)\s*", value):
        hour, minute = map(int, re.findall(r"([0-2]?\d)[:：]([0-5]\d)", value)[0])
        if hour < 24:
            return [hour * 60 + minute]
        
    return []

def extract_stop_times_from_text(text, stop_name):
    times = []
    for match in re.finditer(re.escape(stop_name), text):
        window = text[match.end():match.end() + 24]
        time_match = re.search(r"\(?\s*([0-2]?\d)[:：]([0-5]\d)", window)
        if time_match:
            hour = int(time_match.group(1))
            minute = int(time_match.group(2))
            if hour < 24:
                times.append(hour * 60 + minute)
    return times


def format_minutes_to_time(value):
    if value is None:
        return ""
    return f"{value // 60:02d}:{value % 60:02d}"


def collect_bus_times_from_source(source):
    workbook_path = RAW_DIR / source["file"]
    if not workbook_path.exists():
        return [], set(), f"missing file: {source['file']}"

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if source["sheet"] not in workbook.sheetnames:
        return [], set(), f"missing sheet: {source['sheet']}"

    worksheet = workbook[source["sheet"]]
    direction_keyword = source.get("direction_keyword")
    stop_names = source.get("stop_names", [])
    stop_keys = [normalize_text(stop) for stop in stop_names]
    route_numbers = set(source.get("route_numbers", []))
    note_keywords = source.get("note_keywords", [])

    active_section = direction_keyword is None
    headers = []
    times = []
    routes = set()

    for row in worksheet.iter_rows(values_only=True):
        values = [value for value in row if value is not None]
        if not values:
            continue

        if len(values) == 1 and isinstance(values[0], str):
            active_section = direction_keyword is None or direction_keyword in values[0]
            headers = []
            continue

        if any(normalize_text(value) == "번호" for value in row if value is not None):
            headers = [normalize_text(value) for value in row]
            continue

        if not active_section or not headers or row[0] is None:
            continue

        route_number = str(row[0]).strip()
        if route_numbers and route_number not in route_numbers:
            continue

        row_text = " ".join(str(value) for value in row if value is not None)
        if note_keywords and not any(keyword in row_text for keyword in note_keywords):
            continue

        for idx, header in enumerate(headers):
            if idx >= len(row):
                continue

            if any(stop_key == header for stop_key in stop_keys):
                for minute_value in extract_exact_time(row[idx]):
                    times.append(minute_value)
                    routes.add(route_number)

        for value in row[1:]:
            if not isinstance(value, str):
                continue

            for stop_name in stop_names:
                for minute_value in extract_stop_times_from_text(value, stop_name):
                    times.append(minute_value)
                    routes.add(route_number)

    return sorted(set(times)), routes, ""


def summarize_bus_times(times):
    if len(times) < 2:
        return {
            "observed_runs_count": len(times),
            "bus_interval_min": None,
            "avg_gap_min": None,
            "max_gap_min": None,
            "first_bus_time": format_minutes_to_time(times[0]) if times else "",
            "last_bus_time": format_minutes_to_time(times[-1]) if times else "",
        }

    gaps = [later - earlier for earlier, later in zip(times, times[1:]) if 0 < later - earlier <= 360]

    return {
        "observed_runs_count": len(times),
        "bus_interval_min": int(round(median(gaps))) if gaps else None,
        "avg_gap_min": int(round(sum(gaps) / len(gaps))) if gaps else None,
        "max_gap_min": max(gaps) if gaps else None,
        "first_bus_time": format_minutes_to_time(times[0]),
        "last_bus_time": format_minutes_to_time(times[-1]),
    }


def calculate_bus_schedule_metrics():
    print("\n[3] 버스 엑셀 시간표 기반 관광지별 배차 산출 중...")
    metrics = {}

    for spot_id, rule in BUS_SCHEDULE_RULES.items():
        all_times = []
        all_routes = set()
        notes = []

        for source in rule["sources"]:
            times, routes, error = collect_bus_times_from_source(source)
            all_times.extend(times)
            all_routes.update(routes)
            if error:
                notes.append(error)

        all_times = sorted(set(all_times))
        summary = summarize_bus_times(all_times)

        used_fallback = summary["bus_interval_min"] is None
        interval = summary["bus_interval_min"] or rule["fallback_interval_min"]
        routes_count = len(all_routes) or rule["fallback_routes_count"]

        if used_fallback:
            notes.append(rule.get("fallback_reason", "시간표 자동 산출값 부족으로 기존 기준값 유지"))

        metrics[spot_id] = {
            "spot_id": spot_id,
            "bus_stop_basis": rule["label"],
            "bus_interval_min": interval,
            "bus_routes_count": routes_count,
            "observed_runs_count": summary["observed_runs_count"],
            "avg_gap_min": summary["avg_gap_min"],
            "max_gap_min": summary["max_gap_min"],
            "first_bus_time": summary["first_bus_time"],
            "last_bus_time": summary["last_bus_time"],
            "bus_route_numbers": ", ".join(sorted(all_routes)),
            "bus_schedule_source": "fallback" if used_fallback else "excel_auto",
            "bus_schedule_note": " / ".join(notes),
        }

        print(
            f"  - {spot_id}: {interval}분, "
            f"{routes_count}개 노선, {summary['observed_runs_count']}회 운행 "
            f"({metrics[spot_id]['bus_schedule_source']})"
        )

    return metrics

def build_final_dataset(df_spots_rank, town_entropy, bus_metrics):
    """실측 버스 시간표 팩트와 데이터랩 순위를 결합하여 최종 데이터셋 구축"""
    print("\n[3] 거제시 10대 핵심 거점 실측 지표 매핑 중...")

    # 거제시 핵심 거점 정보 (실측 버스 엑셀 시간표 팩트 반영)
    spots_master = [
        {
            "id": "gohyeon_terminal",
            "name": "고현시외버스터미널 (도심 거점)",
            "category": "hub",
            "lat": 34.8893,
            "lng": 128.6225,
            "town": "고현동",
            "bus_interval_min": 10,
            "bus_routes_count": 24,
            "walk_distance_m": 50,
            "description": "거제 대중교통의 출발 허브. 시내버스 24개 노선 집중 및 도심 생활 상권."
        },
        {
            "id": "wind_hill",
            "name": "바람의 언덕 / 신선대 (남부면)",
            "category": "attraction",
            "lat": 34.7618,
            "lng": 128.6247,
            "town": "남부면",
            "bus_interval_min": 120,  # 55번 실측 120분
            "bus_routes_count": 1,
            "walk_distance_m": 650,
            "description": "외지인 Tmap 검색 1위 대표 명소. 55번 버스 배차 120분으로 극심한 교통 고립지."
        },
        {
            "id": "maemi_castle",
            "name": "매미성 (장목면)",
            "category": "attraction",
            "lat": 34.9818,
            "lng": 128.7186,
            "town": "장목면",
            "bus_interval_min": 95,   # 직통 부재, 환승 대기 포함
            "bus_routes_count": 1,
            "walk_distance_m": 450,
            "description": "외지인 Tmap 검색 2위 포토스팟. 직통 노선 부족으로 환승 필수인 교통 취약지."
        },
        {
            "id": "hakdong_pebble",
            "name": "학동 흑진주 몽돌해변 (동부면)",
            "category": "attraction",
            "lat": 34.7865,
            "lng": 128.6368,
            "town": "동부면",
            "bus_interval_min": 80,
            "bus_routes_count": 2,
            "walk_distance_m": 120,
            "description": "외지인 Tmap 검색 7위 자연경관. 남부 관광의 길목이나 배차 80분으로 접근성 저하."
        },
        {
            "id": "geoje_panoramic",
            "name": "거제 파노라마 케이블카 (동부면)",
            "category": "attraction",
            "lat": 34.7981,
            "lng": 128.6189,
            "town": "동부면",
            "bus_interval_min": 90,
            "bus_routes_count": 1,
            "walk_distance_m": 500,
            "description": "외지인 Tmap 검색 8위 레저시설. 노자산 산간 지대에 위치해 자가용 의존도 95% 이상."
        },
        {
            "id": "gohyeon_market",
            "name": "거제 고현시장 (도심 전통시장)",
            "category": "attraction",
            "lat": 34.8872,
            "lng": 128.6258,
            "town": "고현동",
            "bus_interval_min": 10,
            "bus_routes_count": 18,
            "walk_distance_m": 80,
            "description": "외지인 Tmap 검색 5위 전통시장. 도심지에 위치하여 대중교통 접근성 및 상권 다양성 우수."
        },
        {
            "id": "geoje_jungle_dome",
            "name": "거제 정글돔 / 식물원 (거제면)",
            "category": "attraction",
            "lat": 34.8585,
            "lng": 128.5833,
            "town": "거제면",
            "bus_interval_min": 110,
            "bus_routes_count": 1,
            "walk_distance_m": 250,
            "description": "외지인 Tmap 검색 6위 실내 테마공원. 가족 단위 방문객 급증하나 외곽 배차 110분."
        }
    ]

    for spot in spots_master:
        metric = bus_metrics.get(spot["id"])
        if not metric:
            continue

        spot["bus_interval_min"] = metric["bus_interval_min"]
        spot["bus_routes_count"] = metric["bus_routes_count"]
        spot["bus_stop_basis"] = metric["bus_stop_basis"]
        spot["observed_runs_count"] = metric["observed_runs_count"]
        spot["avg_gap_min"] = metric["avg_gap_min"]
        spot["max_gap_min"] = metric["max_gap_min"]
        spot["first_bus_time"] = metric["first_bus_time"]
        spot["last_bus_time"] = metric["last_bus_time"]
        spot["bus_route_numbers"] = metric["bus_route_numbers"]
        spot["bus_schedule_source"] = metric["bus_schedule_source"]
        spot["bus_schedule_note"] = metric["bus_schedule_note"]

    records = []
    for spot in spots_master:
        # 데이터랩 순위 매칭
        match_rank = df_spots_rank[df_spots_rank["관광지명"].str.contains(spot["name"].split()[0], na=False)]
        tmap_rank = int(match_rank.iloc[0]["순위"]) if not match_rank.empty else 10
        
        # 월간 추정 방문객 (순위에 따른 데이터랩 규모 매핑)
        visitors_monthly = int(120000 / np.sqrt(tmap_rank)) if spot["category"] != "hub" else 195000

        # 실제 상권 엔트로피(CEI) 매핑
        cei_score = town_entropy.get(spot["town"], 0.50)

        # 실제 교통 고립지 지수(TII) 계산
        # - 수요: 월간 방문객이 많고 Tmap 순위가 높을수록 교통 개선 필요성이 커짐
        # - 공급: 버스 노선 수, 배차 간격, 정류장-관광지 도보거리를 함께 반영
        rank_weight = 1 / np.log2(tmap_rank + 2)
        tii_demand_pressure = visitors_monthly * rank_weight

        bus_frequency_per_hour = (spot["bus_routes_count"] * 60) / max(spot["bus_interval_min"], 1)
        last_mile_access_index = 1000 / max(spot["walk_distance_m"], 50)
        tii_supply_index = bus_frequency_per_hour * last_mile_access_index

        tii_raw_score = tii_demand_pressure / (tii_supply_index + 1e-6)
        tii_formula_basis = "TII = log1p(관광수요압력 ÷ 대중교통공급지수) 정규화"
        records.append({
            **spot,
            "visitors_monthly": visitors_monthly,
            "tmap_search_rank": tmap_rank,
            "cei_score": cei_score,
            "tii_demand_pressure": round(float(tii_demand_pressure), 2),
            "bus_frequency_per_hour": round(float(bus_frequency_per_hour), 2),
            "last_mile_access_index": round(float(last_mile_access_index), 2),
            "tii_supply_index": round(float(tii_supply_index), 2),
            "tii_raw_score": round(float(tii_raw_score), 2),
            "tii_formula_basis": tii_formula_basis,
        })

    # for문이 끝난 후 DataFrame 생성 (들여쓰기 정상화)
    df_final = pd.DataFrame(records)

    # TII 원점수는 관광지별 편차가 커서 log1p 후 관광지 구간만 0.20~0.95로 정규화
    non_hub_mask = df_final["category"] != "hub"
    df_final["tii_log_score"] = np.log1p(df_final["tii_raw_score"])

    min_t = df_final.loc[non_hub_mask, "tii_log_score"].min()
    max_t = df_final.loc[non_hub_mask, "tii_log_score"].max()

    if max_t == min_t:
        df_final.loc[non_hub_mask, "tii_score"] = 0.55
    else:
        df_final.loc[non_hub_mask, "tii_score"] = np.round(
            0.20 + (df_final.loc[non_hub_mask, "tii_log_score"] - min_t) / (max_t - min_t) * 0.75,
            2
        )

    df_final.loc[df_final["category"] == "hub", "tii_score"] = 0.12
    # 상태 분류
    df_final["status"] = df_final.apply(
        lambda r: "hub" if r["category"] == "hub" else ("isolated" if r["tii_score"] >= 0.65 else "normal"),
        axis=1
    )
    
    df_final = df_final.drop(columns=["town"])
    return df_final


def main():
    df_spots_rank, avg_stay_geoje, df_geoje_stores = load_raw_datasets()
    town_entropy = calculate_entropy_by_town(df_geoje_stores)
    bus_metrics = calculate_bus_schedule_metrics()
    df_final = build_final_dataset(df_spots_rank, town_entropy, bus_metrics)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / "geoje_real_spots.csv"
    df_final.to_csv(out_path, index=False, encoding="utf-8-sig")

    bus_summary_path = PROCESSED_DIR / "bus_schedule_summary.csv"
    pd.DataFrame(bus_metrics.values()).to_csv(bus_summary_path, index=False, encoding="utf-8-sig")
    print(f"[완료] 버스 시간표 산출 근거 저장 위치: {bus_summary_path}")
    print(f"\n[완료] 실제 데이터 가공 완료! 저장 위치: {out_path}")
    print("\n--- [결과] 거제시 관광 거점 실측 분석 결과 ---")
    print(df_final[["name", "tmap_search_rank", "bus_interval_min", "tii_score", "cei_score", "status"]].to_string(index=False))


if __name__ == "__main__":
    main()
